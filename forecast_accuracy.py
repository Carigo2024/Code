"""
Motor de acuracidade do forecast de importacao (UPI) - Etapa 5.

Ingestao parametrizada dos arquivos Order_NN + Received, construcao do grao
Item x Mes-alvo x Safra, calculo das metricas do dicionario aprovado (Etapa 3)
e exportacao das tabelas (CSV + Excel Power Pivot-ready) e do JSON que alimenta
o dashboard HTML (Etapa 6).

Regras-chave (aprovadas nas Etapas 1-3):
  * Safra = numero do arquivo (Order_NN -> mes N/2026); Order_03 comeca em Abr.
  * Recebido = fornecedor UPI, coluna 'Qtd Delivered', mes de entrada fisica.
  * Chave do item = COMPONENT UDB normalizado (sem '-BR', sem traco final).
  * Leitura honesta = lag fixo N=2 (lead time de importacao ~2 meses).
  * Defasagem fisica sistematica ~ +1 mes (recebido chega depois do previsto).
  * Mes corrente parcial (Julho/26) fica fora dos calculos de acuracidade.
  * Toda % agregada e ponderada por volume; nunca media simples.

Uso:
    python forecast_accuracy.py                 # gera CSV + Excel + JSON
    python forecast_accuracy.py --lag 2 --outdir saida
"""

from __future__ import annotations

import argparse
import datetime as dt
import glob
import json
import os
import re
from dataclasses import dataclass, field, asdict

import numpy as np
import pandas as pd


# --------------------------------------------------------------------------- #
# Configuracao (parametros aprovados)
# --------------------------------------------------------------------------- #
@dataclass
class Config:
    data_dir: str = "data"
    fornecedor: str = "UPI"           # escopo intercompany
    lag_honesto: int = 2              # N do lag fixo
    offset_defasagem: int = 1         # recebido chega ~1 mes depois
    mes_corrente_parcial: str = "2026-07"   # excluido da acuracidade
    # limiares da classificacao de padrao de erro (fracoes)
    t_controle_acum: float = 0.15
    t_controle_mes: float = 0.35
    t_bias: float = 0.20
    t_timing_acum: float = 0.25
    t_timing_mes: float = 0.35
    abc_a: float = 0.80               # corte curva A (% acumulado do volume)
    abc_b: float = 0.95               # corte curva B
    logo_path: str = "data/logo.png"  # logo da empresa (se existir, entra nas abas)


# --------------------------------------------------------------------------- #
# Utilidades de tempo e item
# --------------------------------------------------------------------------- #
def mk(y, m):
    return y * 12 + (m - 1)


def yl(k):
    return f"{k // 12:04d}-{k % 12 + 1:02d}"


def qk(k):
    y, m = k // 12, k % 12 + 1
    return f"{y}Q{(m - 1) // 3 + 1}"


def norm_item(x):
    s = str(x).strip().upper()
    s = re.sub(r"-BR$", "", s)
    s = re.sub(r"-$", "", s)
    return s


def parse_header_month(h):
    """Chave de mes de um cabecalho de coluna, tratando a corrupcao do Order_06.

    '01/05/20267' (texto) e interpretado como 2027-05 (dia 01, mes 05).
    """
    if isinstance(h, dt.datetime):
        return mk(h.year, h.month)
    if isinstance(h, str):
        m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})\d?\s*$", h)
        if m:
            _, mes, _ = m.groups()
            return mk(2027, int(mes))
    return None


# --------------------------------------------------------------------------- #
# Ingestao
# --------------------------------------------------------------------------- #
def _mes_do_order(base):
    """Mes (1-12) de um nome 'Order 01 2026.xlsx' ou 'Order_01_2026.xlsx'; senao None."""
    if base.startswith("~"):
        return None
    m = re.search(r"Order[ _\-]*(\d{1,2})", base, re.IGNORECASE)
    if not m:
        return None
    n = int(m.group(1))
    return n if 1 <= n <= 12 else None


def _lista_orders(data_dir):
    """Arquivos Order (aceita espaco ou underscore no nome), ordenados por mes."""
    out = []
    for c in glob.glob(os.path.join(data_dir, "Order*.xlsx")):
        n = _mes_do_order(os.path.basename(c))
        if n:
            out.append((c, n))
    return sorted(out, key=lambda t: t[1])


def _arquivo_recebido(data_dir):
    achados = [c for c in sorted(glob.glob(os.path.join(data_dir, "Received*.xlsx")))
               if not os.path.basename(c).startswith("~")]
    if not achados:
        raise FileNotFoundError(f"arquivo Received*.xlsx nao encontrado em {data_dir}")
    return achados[0]


def le_forecasts(cfg: Config):
    """Grao tidy do forecast: item, alvo, safra, lag, fcst, descricao, um."""
    import openpyxl

    recs = []
    desc_map, um_map = {}, {}
    for caminho, _n in _lista_orders(cfg.data_dir):
        safra = mk(2026, _n)
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb["FORECAST"]
        linhas = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        hdr = linhas[0]
        vistos, cols = set(), []
        for j in range(4, len(hdr)):
            tk = parse_header_month(hdr[j])
            if tk is None or tk in vistos:      # dedupe Order_03 (Fev/27 repetido)
                continue
            vistos.add(tk)
            cols.append((j, tk))
        for r in linhas[1:]:
            if r[0] in (None, "") or not str(r[0]).strip():
                continue
            it = norm_item(r[0])
            if r[2] not in (None, ""):
                desc_map.setdefault(it, str(r[2]).strip())
            if r[3] not in (None, ""):
                um_map.setdefault(it, str(r[3]).strip())
            for j, tk in cols:
                if r[j] is None:
                    continue
                recs.append((it, tk, safra, tk - safra, float(r[j])))
    fc = pd.DataFrame(recs, columns=["item", "alvo", "safra", "lag", "fcst"])
    return fc, desc_map, um_map


def le_recebido(cfg: Config):
    """Recebido UPI agregado por item x mes de entrada; e dimensoes do item."""
    import openpyxl

    caminho = _arquivo_recebido(cfg.data_dir)
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    linhas = list(wb["Export"].iter_rows(min_row=2, values_only=True))
    wb.close()

    recs, dims = [], {}
    descartadas = 0
    for r in linhas:
        if r[5] in (None, ""):
            continue
        if str(r[4]).strip() != cfg.fornecedor:
            continue
        d, q = r[1], r[12]
        if not isinstance(d, dt.datetime) or q is None:
            descartadas += 1
            continue
        it = norm_item(r[5])
        recs.append((it, mk(d.year, d.month), float(q)))
        dd = dims.setdefault(it, {"desc": None, "familia": None, "marca": None})
        if dd["familia"] is None and r[8] not in (None, ""):
            dd["familia"] = str(r[8]).strip()
        if dd["marca"] is None and r[9] not in (None, ""):
            dd["marca"] = str(r[9]).strip()
        if dd["desc"] is None and r[6] not in (None, ""):
            dd["desc"] = str(r[6]).strip()
    act = (
        pd.DataFrame(recs, columns=["item", "alvo", "recv"])
        .groupby(["item", "alvo"], as_index=False)
        .recv.sum()
    )
    return act, dims, descartadas


# --------------------------------------------------------------------------- #
# Selecao das previsoes (modos de leitura)
# --------------------------------------------------------------------------- #
def sel_lag(fc, n):
    return fc[fc.lag == n][["item", "alvo", "fcst"]].rename(columns={"fcst": "F"})


def sel_consolidado(fc):
    return (
        fc.sort_values("safra")
        .groupby(["item", "alvo"])
        .tail(1)[["item", "alvo", "fcst"]]
        .rename(columns={"fcst": "F"})
    )


# --------------------------------------------------------------------------- #
# Metricas
# --------------------------------------------------------------------------- #
def malha(sel, act, janela):
    """Junta previsto x recebido na janela (outer, zera faltantes)."""
    m = sel.merge(act.rename(columns={"recv": "A"}), on=["item", "alvo"], how="outer")
    m = m[m.alvo.isin(janela)].copy()
    m[["F", "A"]] = m[["F", "A"]].fillna(0.0)
    m["e"] = m.A - m.F
    return m


def wape(m, col_f="F", col_a="A"):
    A = m[col_a].sum()
    return float((m[col_f] - m[col_a]).abs().sum() / A) if A else np.nan


def metricas_agregadas(m):
    """Bias e WAPE nas tres leituras de timing, para um recorte ja filtrado."""
    F, A = m.F.sum(), m.A.sum()
    bias = (A - F) / F if F else np.nan
    wm = wape(m)
    g = m.assign(t=m.alvo.map(qk)).groupby(["item", "t"]).agg(F=("F", "sum"), A=("A", "sum"))
    wq = float((g.F - g.A).abs().sum() / A) if A else np.nan
    gc = m.groupby("item").agg(F=("F", "sum"), A=("A", "sum"))
    wc = float((gc.F - gc.A).abs().sum() / A) if A else np.nan
    return dict(F=float(F), A=float(A), bias=bias, wape_mes=wm, wape_trim=wq,
                wape_acum=wc, acuracia=(1 - wc) if not np.isnan(wc) else np.nan,
                n_itens=int(m.item.nunique()))


def tracking_signal(serie_e):
    """TS = soma corrente do erro / MAD."""
    e = np.asarray(serie_e, float)
    mad = np.mean(np.abs(e))
    return float(np.sum(e) / mad) if mad else 0.0


def churn_por_item(fc, closed_max):
    """Volatilidade da previsao de um mesmo mes-alvo entre safras sucessivas."""
    out = {}
    sub = fc[fc.alvo <= closed_max + 6]     # inclui horizonte proximo
    for it, g in sub.groupby("item"):
        num = den = 0.0
        for T, gt in g.sort_values("safra").groupby("alvo"):
            vals = gt.fcst.values
            if len(vals) >= 2:
                num += np.abs(np.diff(vals)).sum()
                den += np.mean(vals)
        out[it] = float(num / den) if den else 0.0
    return out


# --------------------------------------------------------------------------- #
# Classificacao de padrao de erro (limiares aprovados)
# --------------------------------------------------------------------------- #
def classifica(bias, wape_mes, wape_acum, cfg: Config, tem_forecast):
    if not tem_forecast:
        return "nao_previsto"
    if np.isnan(bias):
        return "nao_previsto"
    if wape_acum <= cfg.t_controle_acum and wape_mes <= cfg.t_controle_mes:
        return "sob_controle"
    if bias <= -cfg.t_bias:
        return "vies_super"
    if bias >= cfg.t_bias:
        return "vies_sub"
    if abs(bias) < cfg.t_bias and wape_acum <= cfg.t_timing_acum and wape_mes >= cfg.t_timing_mes:
        return "timing"
    return "erratico"


ACAO = {
    "sob_controle": "Nenhuma - dentro da tolerancia.",
    "vies_super": "Superprevisao cronica: corrigir metodo/premissa; revisar estoque (capital parado).",
    "vies_sub": "Subprevisao cronica: revisar premissa e estoque de seguranca (risco de ruptura).",
    "timing": "Volume bate, meses nao: tratar com lead time (+1 mes) e bucket, nao mudar a previsao.",
    "erratico": "Alta dispersao sem vies: cobrir com estoque de seguranca, nao adianta prever melhor.",
    "nao_previsto": "Recebido sem previsao na janela: incluir no forecast da matriz.",
}
CLASSE_LABEL = {
    "sob_controle": "Sob controle", "vies_super": "Vies superprevisao",
    "vies_sub": "Vies subprevisao", "timing": "Erro de timing",
    "erratico": "Erratico", "nao_previsto": "Nao previsto",
}


# --------------------------------------------------------------------------- #
# Construcao do modelo completo
# --------------------------------------------------------------------------- #
def build_model(cfg: Config):
    fc, desc_fc, um_map = le_forecasts(cfg)
    act, dims, descartadas = le_recebido(cfg)

    parcial = mk(*map(int, cfg.mes_corrente_parcial.split("-")))
    closed_max = parcial - 1
    meses_fechados = sorted(t for t in set(act.alvo) if t <= closed_max)

    overlap = sorted(set(fc.item) & set(act.item))
    fc_o = fc[fc.item.isin(overlap)]
    act_o = act[act.item.isin(overlap) & (act.alvo <= closed_max)]

    # ---- janelas ----
    win_lag = sorted(t for t in set(sel_lag(fc_o, cfg.lag_honesto).alvo) & set(act_o.alvo)
                     if t <= closed_max)
    win_full = meses_fechados

    # ---- metricas agregadas (headline) ----
    agg = {
        "lag_honesto": metricas_agregadas(malha(sel_lag(fc_o, cfg.lag_honesto), act_o, win_lag)),
        "consolidado_win": metricas_agregadas(malha(sel_consolidado(fc_o), act_o, win_lag)),
        "consolidado_full": metricas_agregadas(malha(sel_consolidado(fc_o), act_o, win_full)),
    }

    # ---- sensibilidade por lag ----
    sens = []
    for L in range(0, 12):
        w = sorted(t for t in set(sel_lag(fc_o, L).alvo) & set(act_o.alvo) if t <= closed_max)
        if not w:
            continue
        r = metricas_agregadas(malha(sel_lag(fc_o, L), act_o, w))
        sens.append(dict(lag=L, janela=[yl(t) for t in w], bias=r["bias"],
                         wape_mes=r["wape_mes"], wape_acum=r["wape_acum"]))

    # ---- grade de cobertura ----
    cobertura = []
    for T in meses_fechados:
        lags = sorted(set(fc_o[fc_o.alvo == T].lag))
        cobertura.append(dict(alvo=yl(T), lags=lags, tem_realizado=T in set(act_o.alvo)))

    # ---- churn ----
    churn = churn_por_item(fc_o, closed_max)

    # ---- serie mensal por item (leitura honesta lag N) + metricas por item ----
    mL = malha(sel_lag(fc_o, cfg.lag_honesto), act_o, win_lag)
    volpi = act_o[act_o.alvo.isin(win_lag)].groupby("item").recv.sum()
    # forecast presence for status (nonzero em V01 vs V07)
    v_min, v_max = fc.safra.min(), fc.safra.max()
    pres = (fc[fc.fcst > 0].groupby("item").safra
            .agg(["min", "max"]).rename(columns={"min": "vmin", "max": "vmax"}))

    itens = []
    series = {}
    for it in overlap:
        g = mL[mL.item == it].sort_values("alvo")
        A = g.A.sum(); F = g.F.sum()
        if len(g) == 0:
            continue
        bias = (A - F) / F if F > 0 else np.nan
        wm = wape(g); wc = abs(A - F) / A if A > 0 else np.nan
        ts = tracking_signal(g.e.values)
        tem_fc = F > 0
        classe = classifica(bias, wm, wc, cfg, tem_fc)
        # status estrutural
        pr = pres.loc[it] if it in pres.index else None
        if pr is None:
            status = "sem_forecast"
        elif pr.vmin > v_min and pr.vmax >= v_max:
            status = "novo"
        elif pr.vmax < v_max:
            status = "descontinuado"
        else:
            status = "ativo"
        dd = dims.get(it, {})
        itens.append(dict(
            item=it,
            desc=(dd.get("desc") or desc_fc.get(it) or ""),
            familia=dd.get("familia") or "",
            marca=dd.get("marca") or "",
            um=um_map.get(it, ""),
            vol_recv=float(volpi.get(it, 0.0)),
            F_lag=float(F), A=float(A),
            bias=None if np.isnan(bias) else round(bias, 4),
            wape_mes=None if np.isnan(wm) else round(wm, 4),
            wape_acum=None if np.isnan(wc) else round(wc, 4),
            acuracia=None if np.isnan(wc) else round(1 - wc, 4),
            tracking_signal=round(ts, 2),
            churn=round(churn.get(it, 0.0), 4),
            n_meses=int((g.A > 0).sum()),
            classe=classe, classe_label=CLASSE_LABEL[classe], acao=ACAO[classe],
            status=status,
        ))
        series[it] = dict(
            meses=[yl(t) for t in g.alvo],
            F=[round(x, 1) for x in g.F], A=[round(x, 1) for x in g.A],
        )
    mestre = pd.DataFrame(itens)

    # ---- ABC por volume recebido ----
    mestre = mestre.sort_values("vol_recv", ascending=False).reset_index(drop=True)
    tot = mestre.vol_recv.sum()
    mestre["cum_pct"] = mestre.vol_recv.cumsum() / tot if tot else 0.0
    mestre["abc"] = np.where(mestre.cum_pct <= cfg.abc_a, "A",
                     np.where(mestre.cum_pct <= cfg.abc_b, "B", "C"))

    # ---- rankings (4) ----
    ok = mestre[mestre.A > 0].copy()
    ok["erro_abs_un"] = (ok.A - ok.F_lag).abs()
    top_abs_over = ok[ok.F_lag > ok.A].nlargest(15, "erro_abs_un")
    top_abs_under = ok[ok.A > ok.F_lag].nlargest(15, "erro_abs_un")
    ab = ok[ok.abc.isin(["A", "B"])].copy()          # piso de volume
    top_pct_over = ab[ab.bias < 0].nsmallest(15, "bias")
    top_pct_under = ab[ab.bias > 0].nlargest(15, "bias")

    # ---- views por lag (para o seletor interativo do dashboard) ----
    itemdim = {r["item"]: dict(desc=r["desc"], familia=r["familia"], marca=r["marca"],
                               um=r["um"], abc=r["abc"], status=r["status"],
                               churn=r["churn"], vol_recv=r["vol_recv"])
               for r in mestre.to_dict("records")}
    abc_map = dict(zip(mestre.item, mestre.abc))

    def _rankings_view(recs):
        df = pd.DataFrame([r for r in recs if r["A"] and r["A"] > 0])
        if df.empty:
            return dict(abs_over=[], abs_under=[], pct_over=[], pct_under=[])
        df["abc"] = df.item.map(lambda i: abc_map.get(i, "C"))
        df["erro_abs_un"] = (df.A - df.F_lag).abs()
        cols = ["item", "abc", "F_lag", "A", "bias", "wape_acum", "erro_abs_un"]
        rl = lambda d: d[cols].to_dict("records")
        ab = df[df.abc.isin(["A", "B"]) & df.bias.notna()]
        return dict(
            abs_over=rl(df[df.F_lag > df.A].nlargest(15, "erro_abs_un")),
            abs_under=rl(df[df.A > df.F_lag].nlargest(15, "erro_abs_un")),
            pct_over=rl(ab[ab.bias < 0].nsmallest(15, "bias")),
            pct_under=rl(ab[ab.bias > 0].nlargest(15, "bias")))

    def _view(sel, win):
        m = malha(sel, act_o, win)
        recs = []
        for it, g in m.groupby("item"):
            g = g.sort_values("alvo")
            A, F = g.A.sum(), g.F.sum()
            if A <= 0 and F <= 0:
                continue
            bias = (A - F) / F if F > 0 else np.nan
            wm = wape(g); wc = abs(A - F) / A if A > 0 else np.nan
            classe = classifica(bias, wm, wc, cfg, F > 0)
            recs.append(dict(
                item=it, F_lag=float(F), A=float(A),
                bias=None if np.isnan(bias) else round(bias, 4),
                wape_mes=None if np.isnan(wm) else round(wm, 4),
                wape_acum=None if np.isnan(wc) else round(wc, 4),
                acuracia=None if np.isnan(wc) else round(1 - wc, 4),
                tracking_signal=round(tracking_signal(g.e.values), 2),
                n_meses=int((g.A > 0).sum()),
                classe=classe, classe_label=CLASSE_LABEL[classe], acao=ACAO[classe]))
        A_m = act_o[act_o.alvo.isin(win)].groupby("alvo").recv.sum()
        F_m = sel[sel.alvo.isin(win)].groupby("alvo").F.sum()
        meses = sorted(win)
        ov = dict(meses=[yl(t) for t in meses],
                  F=[round(float(F_m.get(t, 0)), 1) for t in meses],
                  A=[round(float(A_m.get(t, 0)), 1) for t in meses])
        ov["err"] = [None if not ov["F"][i] else round((ov["A"][i] - ov["F"][i]) / ov["F"][i] * 100, 1)
                     for i in range(len(meses))]
        return dict(agg=metricas_agregadas(m), metrics=recs, overview=ov,
                    rankings=_rankings_view(recs), window=[yl(t) for t in meses])

    views, lags_incluidos = {}, []
    for L in range(0, 6):
        sel = sel_lag(fc_o, L)
        w = sorted(t for t in set(sel.alvo) & set(act_o.alvo) if t <= closed_max)
        if len(w) < 2:
            continue
        views[str(L)] = _view(sel, w)
        lags_incluidos.append(L)
    views["cons"] = _view(sel_consolidado(fc_o), win_full)

    # ---- orfaos ----
    prev_nao_receb = sorted(set(fc[fc.fcst > 0].item) - set(act.item))
    receb_sem_prev = sorted(set(act.item) - set(fc[fc.fcst > 0].item))

    return dict(
        cfg=cfg, fc=fc, fc_o=fc_o, act=act, act_o=act_o,
        overlap=overlap, meses_fechados=meses_fechados,
        win_lag=win_lag, win_full=win_full, closed_max=closed_max,
        agg=agg, sensibilidade=sens, cobertura=cobertura,
        mestre=mestre, series=series,
        views=views, itemdim=itemdim, lags_incluidos=lags_incluidos,
        rankings=dict(abs_over=top_abs_over, abs_under=top_abs_under,
                      pct_over=top_pct_over, pct_under=top_pct_under),
        orfaos=dict(previsto_nao_recebido=prev_nao_receb, recebido_sem_previsao=receb_sem_prev),
        descartadas=descartadas,
    )


# --------------------------------------------------------------------------- #
# Exportacao
# --------------------------------------------------------------------------- #
def exporta_csv(model, outdir):
    os.makedirs(outdir, exist_ok=True)
    model["fc"].assign(alvo=model["fc"].alvo.map(yl), safra=model["fc"].safra.map(yl)) \
        .to_csv(os.path.join(outdir, "grao_forecast.csv"), index=False)
    model["act"].assign(alvo=model["act"].alvo.map(yl)) \
        .to_csv(os.path.join(outdir, "recebido_upi.csv"), index=False)
    model["mestre"].drop(columns=[c for c in ["cum_pct"] if c in model["mestre"].columns]) \
        .to_csv(os.path.join(outdir, "mestre_itens.csv"), index=False)
    # metricas por item x lag (tidy) - ideal para Tabela Dinamica + segmentacao de 'lag'
    plag = []
    for k, v in model["views"].items():
        lagname = "consolidado" if k == "cons" else f"lag {k}"
        for m in v["metrics"]:
            plag.append(dict(lag=lagname, item=m["item"], previsto=m["F_lag"], recebido=m["A"],
                             bias=m["bias"], wape_mes=m["wape_mes"], wape_acum=m["wape_acum"],
                             acuracia=m["acuracia"], tracking_signal=m["tracking_signal"],
                             classe=m["classe_label"]))
    pd.DataFrame(plag).to_csv(os.path.join(outdir, "metricas_por_lag.csv"), index=False)


def exporta_excel(model, caminho):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    cfg = model["cfg"]
    # --- paleta laranja ---
    HDR = PatternFill("solid", fgColor="D9601C")     # cabecalho
    HDR_FONT = Font(bold=True, color="FFFFFF")
    TIT_FONT = Font(bold=True, size=12, color="A8410E")
    BAND = PatternFill("solid", fgColor="FDF3EC")     # faixa zebrada
    LOGO_FILL = PatternFill("solid", fgColor="FCE7DA")
    LOGO_FONT = Font(bold=True, color="B5673C", size=11)
    ord_side = Side(style="thin", color="E2611C")
    LOGO_BORDER = Border(left=ord_side, right=ord_side, top=ord_side, bottom=ord_side)

    tem_logo = os.path.isfile(cfg.logo_path)
    BRANCO = PatternFill("solid", fgColor="FFFFFF")

    def caixa_logo(ws):
        """Reserva o canto superior esquerdo para o logo (imagem, se existir)."""
        ws.merge_cells("A1:C6")
        c = ws["A1"]
        c.fill = BRANCO if tem_logo else LOGO_FILL
        c.value = None if tem_logo else "LOGO DA EMPRESA\n(inserir aqui)"
        c.font = LOGO_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws["A1:C6"]:
            for cell in row:
                cell.border = LOGO_BORDER
                if not tem_logo and cell.fill.patternType is None:
                    cell.fill = LOGO_FILL
        for r in range(1, 7):
            ws.row_dimensions[r].height = 17
        if tem_logo:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(cfg.logo_path)
            alt = 116
            img.height, img.width = alt, int(alt * img.width / img.height)
            img.anchor = "A1"
            ws.add_image(img)
        return 8  # conteudo comeca aqui

    def escreve(ws, df, titulo=None, logo=False, banding=False, fmt=None):
        fmt = fmt or {}
        cols = list(df.columns)
        start = caixa_logo(ws) if logo else 1
        r0 = start
        if titulo:
            ws.cell(r0, 1, titulo).font = TIT_FONT
            r0 = start + 2
        for j, c in enumerate(cols, 1):
            cell = ws.cell(r0, j, c)
            cell.fill = HDR; cell.font = HDR_FONT
            cell.alignment = Alignment(vertical="center")
        for i, row in enumerate(df.itertuples(index=False), r0 + 1):
            zebra = banding and (i - r0) % 2 == 0
            for j, v in enumerate(row, 1):
                cell = ws.cell(i, j, v)
                if zebra:
                    cell.fill = BAND
                nf = fmt.get(cols[j - 1])
                if nf and isinstance(v, (int, float)):
                    cell.number_format = nf
        ws.freeze_panes = ws.cell(r0 + 1, 1)
        ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(cols))}{r0+len(df)}"
        for j, c in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(j)].width = min(28, max(10, len(str(c)) + 3))

    # ---- aba PAINEL (dashboard de uma pagina, abre primeiro) ----
    from openpyxl.chart import BarChart, Reference
    wsp = wb.active; wsp.title = "Painel"
    wsp.sheet_view.showGridLines = False
    v2 = model["views"][str(cfg.lag_honesto)]
    agg = v2["agg"]; ov = v2["overview"]
    caixa_logo(wsp)
    wsp["E2"] = "Painel de Acuracidade  ·  Forecast de Importacao (UPI)"
    wsp["E2"].font = Font(bold=True, size=15, color="A8410E")
    wsp["E4"] = (f"Leitura honesta: lag {cfg.lag_honesto}   ·   janela {', '.join(v2['window'])}"
                 f"   ·   {len(model['overlap'])} itens comparaveis")
    wsp["E4"].font = Font(size=10, color="52514E")
    # KPIs
    for i, (lbl, val) in enumerate([("ACURACIA (volume)", agg["acuracia"]),
                                    ("BIAS de volume", agg["bias"]),
                                    ("WAPE acumulada", agg["wape_acum"]),
                                    ("WAPE mes (com timing)", agg["wape_mes"])]):
        col = 1 + i * 2
        wsp.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col + 1)
        wsp.merge_cells(start_row=9, start_column=col, end_row=10, end_column=col + 1)
        lc = wsp.cell(8, col, lbl); lc.font = Font(size=9, bold=True, color="8A8984")
        lc.alignment = Alignment(horizontal="center")
        vc = wsp.cell(9, col, val); vc.number_format = "0.0%"
        vc.font = Font(bold=True, size=20, color="C2521A")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        for rr in (8, 9, 10):
            for cc in (col, col + 1):
                wsp.cell(rr, cc).border = LOGO_BORDER
    # tabela por lag
    wsp.cell(12, 1, "Acuracidade por lag").font = TIT_FONT
    for j, h in enumerate(["Lag", "Bias", "WAPE mes", "WAPE acum", "Acuracia"], 1):
        c = wsp.cell(13, j, h); c.fill = HDR; c.font = HDR_FONT
    rr = 14
    for k in [str(i) for i in model["lags_incluidos"]] + ["cons"]:
        a = model["views"][k]["agg"]
        wsp.cell(rr, 1, "consolidado" if k == "cons" else "lag " + k)
        for j, val in enumerate([a["bias"], a["wape_mes"], a["wape_acum"], a["acuracia"]], 2):
            wsp.cell(rr, j, val).number_format = "0.0%"
        if k == str(cfg.lag_honesto):
            for j in range(1, 6):
                wsp.cell(rr, j).fill = BAND
        rr += 1
    # grafico Previsto vs Recebido (dados em colunas ocultas N..P)
    hc = 14
    wsp.cell(1, hc, "mes"); wsp.cell(1, hc + 1, "Previsto"); wsp.cell(1, hc + 2, "Recebido")
    for i, mes in enumerate(ov["meses"], 2):
        wsp.cell(i, hc, mes); wsp.cell(i, hc + 1, ov["F"][i - 2]); wsp.cell(i, hc + 2, ov["A"][i - 2])
    ch = BarChart(); ch.type = "col"; ch.title = "Previsto vs Recebido por mes (un.)"
    ch.height = 7.5; ch.width = 15
    data = Reference(wsp, min_col=hc + 1, max_col=hc + 2, min_row=1, max_row=1 + len(ov["meses"]))
    cats = Reference(wsp, min_col=hc, max_col=hc, min_row=2, max_row=1 + len(ov["meses"]))
    ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
    try:
        ch.series[0].graphicalProperties.solidFill = "1B9E77"
        ch.series[1].graphicalProperties.solidFill = "E2611C"
    except Exception:
        pass
    wsp.add_chart(ch, "G12")
    for c in range(hc, hc + 3):
        wsp.column_dimensions[get_column_letter(c)].hidden = True
    # maiores variacoes (top absolutas)
    rk = v2["rankings"]

    def _top(titulo, linhas, startrow):
        wsp.cell(startrow, 1, titulo).font = TIT_FONT
        for j, h in enumerate(["Item", "Descricao", "Erro (un)", "Bias"], 1):
            c = wsp.cell(startrow + 1, j, h); c.fill = HDR; c.font = HDR_FONT
        for i, r in enumerate(linhas[:6], startrow + 2):
            dd = model["itemdim"].get(r["item"], {})
            wsp.cell(i, 1, r["item"])
            wsp.cell(i, 2, (dd.get("desc") or "")[:26])
            wsp.cell(i, 3, r["erro_abs_un"]).number_format = "#,##0"
            wsp.cell(i, 4, r["bias"]).number_format = "0.0%"

    _top("Maiores subprevisoes (recebeu mais)", rk["abs_under"], rr + 1)
    _top("Maiores superprevisoes (previu mais)", rk["abs_over"], rr + 11)
    wsp.column_dimensions["A"].width = 13
    wsp.column_dimensions["B"].width = 28
    for cl in "CDEFGHIJ":
        wsp.column_dimensions[cl].width = 12

    # ---- aba leia-me ----
    ws = wb.create_sheet("leia-me")
    logo_start = caixa_logo(ws)
    notas = [
        "ACURACIDADE DO FORECAST DE IMPORTACAO (UPI)",
        "",
        f"Escopo recebido: fornecedor '{cfg.fornecedor}' (intercompany), coluna Qtd Delivered.",
        f"Leitura honesta: lag fixo N={cfg.lag_honesto} (lead time ~2 meses).",
        f"Defasagem fisica sistematica: +{cfg.offset_defasagem} mes (recebido chega depois).",
        f"Mes corrente parcial excluido: {cfg.mes_corrente_parcial}.",
        f"Itens comparaveis (UPI): {len(model['overlap'])}.",
        "",
        "Limiares de classificacao (fracao):",
        f"  sob_controle: WAPE_acum<={cfg.t_controle_acum} e WAPE_mes<={cfg.t_controle_mes}",
        f"  vies: |bias|>={cfg.t_bias}   timing: WAPE_acum<={cfg.t_timing_acum} e WAPE_mes>={cfg.t_timing_mes}",
        "",
        "Abas:",
        "  mestre_itens     - metricas por item na leitura honesta (lag {}).".format(cfg.lag_honesto),
        "  metricas_por_lag - MESMAS metricas por item em CADA lag (0..4 + consolidado);",
        "                     use com uma Segmentacao de Dados no campo 'lag' para escolher a leitura.",
        "  grao_forecast    - tabela-fato Item x Alvo x Safra x LAG x fcst (Power Pivot).",
        "  recebido_upi     - tabela-fato Item x Alvo x recv (Power Pivot).",
        "  cobertura        - grade mes x lag.",
        "",
        "Percentuais ja formatados (bias, WAPE, acuracia, churn). Para Power Pivot:",
        "Dados > Obter Dados > Deste Arquivo, e relacione grao/recebido por 'item'/'alvo'.",
        "O dashboard HTML tem um seletor de Lag equivalente (Lag 0 = mesmo mes ... Consolidado).",
    ]
    for i, t in enumerate(notas, logo_start):
        cell = ws.cell(i, 1, t)
        if i == logo_start:
            cell.font = TIT_FONT
    ws.column_dimensions["A"].width = 95

    FMT_PCT = "0.0%"
    FMT_MIL = "#,##0"
    fmt_mestre = {"vol_recv": FMT_MIL, "F_lag": FMT_MIL, "A": FMT_MIL,
                  "bias": FMT_PCT, "wape_mes": FMT_PCT, "wape_acum": FMT_PCT,
                  "acuracia": FMT_PCT, "tracking_signal": "0.00", "churn": FMT_PCT}
    mestre_show = model["mestre"].drop(columns=[c for c in ["cum_pct"] if c in model["mestre"].columns])
    escreve(wb.create_sheet("mestre_itens"), mestre_show,
            "Tabela mestre de itens", logo=True, banding=True, fmt=fmt_mestre)

    # ---- metricas por lag (tidy: para Segmentacao de Lag no Excel) ----
    plag = []
    for k, v in model["views"].items():
        lagname = "consolidado" if k == "cons" else f"lag {k}"
        for m in v["metrics"]:
            plag.append(dict(lag=lagname, item=m["item"], previsto=m["F_lag"], recebido=m["A"],
                             bias=m["bias"], wape_mes=m["wape_mes"], wape_acum=m["wape_acum"],
                             acuracia=m["acuracia"], tracking_signal=m["tracking_signal"],
                             classe=m["classe_label"]))
    fmt_plag = {"previsto": FMT_MIL, "recebido": FMT_MIL, "bias": FMT_PCT,
                "wape_mes": FMT_PCT, "wape_acum": FMT_PCT, "acuracia": FMT_PCT,
                "tracking_signal": "0.00"}
    escreve(wb.create_sheet("metricas_por_lag"), pd.DataFrame(plag),
            "Metricas por item x lag (use Segmentacao no campo 'lag')", fmt=fmt_plag)

    cob = pd.DataFrame([{"alvo": c["alvo"], "lags": str(c["lags"]),
                         "tem_realizado": c["tem_realizado"]} for c in model["cobertura"]])
    escreve(wb.create_sheet("cobertura"), cob,
            "Grade de cobertura (mes x lag)", logo=True, banding=True)
    gf = model["fc"].assign(alvo=model["fc"].alvo.map(yl), safra=model["fc"].safra.map(yl))
    escreve(wb.create_sheet("grao_forecast"), gf, fmt={"fcst": FMT_MIL})
    ru = model["act"].assign(alvo=model["act"].alvo.map(yl))
    escreve(wb.create_sheet("recebido_upi"), ru, fmt={"recv": FMT_MIL})
    wb.save(caminho)


def exporta_json(model, caminho):
    cfg = model["cfg"]
    payload = dict(
        meta=dict(
            escopo=cfg.fornecedor, lag=cfg.lag_honesto, offset=cfg.offset_defasagem,
            mes_parcial=cfg.mes_corrente_parcial, n_itens=len(model["overlap"]),
            win_lag=[yl(t) for t in model["win_lag"]],
            win_full=[yl(t) for t in model["win_full"]],
            limiares=dict(controle_acum=cfg.t_controle_acum, controle_mes=cfg.t_controle_mes,
                          bias=cfg.t_bias, timing_acum=cfg.t_timing_acum, timing_mes=cfg.t_timing_mes),
        ),
        agg=model["agg"], sensibilidade=model["sensibilidade"], cobertura=model["cobertura"],
        mestre=model["mestre"].where(pd.notnull(model["mestre"]), None).to_dict("records"),
        series=model["series"],
        rankings={k: v.where(pd.notnull(v), None)[
            ["item", "desc", "familia", "abc", "F_lag", "A", "bias", "erro_abs_un" if "abs" in k else "wape_acum"]
        ].to_dict("records") for k, v in model["rankings"].items()},
        orfaos={k: v[:200] for k, v in model["orfaos"].items()},
    )
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, default=float)


# --------------------------------------------------------------------------- #
# Reconciliacao (previa da Etapa 7)
# --------------------------------------------------------------------------- #
def reconcilia(model):
    """Auto-validacao: soma crua dos arquivos vs modelo (roda a cada atualizacao)."""
    import openpyxl

    cfg = model["cfg"]
    # ---- forecast: soma crua de TODAS as celulas de mes menos a col duplicada ----
    raw_all = dup = 0.0
    for caminho, _n in _lista_orders(cfg.data_dir):
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        linhas = list(wb["FORECAST"].iter_rows(min_row=2, values_only=True))
        wb.close()
        hdr = linhas[0]
        seen = set()
        for j in range(4, len(hdr)):
            tk = parse_header_month(hdr[j])
            if tk is None:
                continue
            colsum = sum(r[j] for r in linhas[1:] if isinstance(r[j], (int, float)))
            raw_all += colsum
            if tk in seen:
                dup += colsum
            else:
                seen.add(tk)
    grain = model["fc"].fcst.sum()
    dif_fc = raw_all - dup - grain

    bruto = model["act"][model["act"].alvo <= model["closed_max"]].recv.sum()
    modelo = model["act_o"].recv.sum()
    print("\n--- Reconciliacao (auto-validacao) ---")
    print(f"Forecast: soma crua {raw_all:,.0f} - duplicada {dup:,.0f} = {raw_all-dup:,.0f} "
          f"| grao {grain:,.0f} | DIF {dif_fc:,.0f} {'OK' if abs(dif_fc)<1 else '*** REVISAR'}")
    print(f"Recebido UPI (fechado): bruto {bruto:,.0f} | modelo comparavel {modelo:,.0f} "
          f"| itens recebidos sem forecast: {len(model['orfaos']['recebido_sem_previsao'])}")
    return abs(dif_fc) < 1


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--data-dir", default="data")
    ap.add_argument("--lag", type=int, default=None)
    ap.add_argument("--outdir", default="saida")
    args = ap.parse_args()

    cfg = Config(data_dir=args.data_dir)
    if args.lag is not None:
        cfg.lag_honesto = args.lag

    model = build_model(cfg)
    os.makedirs(args.outdir, exist_ok=True)
    exporta_csv(model, args.outdir)
    exporta_excel(model, os.path.join(args.outdir, "forecast_accuracy.xlsx"))
    exporta_json(model, os.path.join(args.outdir, "dashboard_data.json"))

    a = model["agg"]["lag_honesto"]
    print(f"Safras: {[yl(s) for s in sorted(model['fc'].safra.unique())]}")
    print(f"Itens comparaveis (UPI): {len(model['overlap'])} | "
          f"janela honesta lag-{cfg.lag_honesto}: {[yl(t) for t in model['win_lag']]}")
    print(f"Headline lag-{cfg.lag_honesto}: bias={a['bias']*100:+.1f}%  "
          f"WAPE mes={a['wape_mes']*100:.1f}% trim={a['wape_trim']*100:.1f}% acum={a['wape_acum']*100:.1f}%")
    vc = model["mestre"].classe_label.value_counts()
    print("Classificacao de itens:\n" + vc.to_string())
    print("ABC:\n" + model["mestre"].abc.value_counts().reindex(["A", "B", "C"]).to_string())
    reconcilia(model)
    print(f"\nSaidas em: {args.outdir}/  (CSV x3, forecast_accuracy.xlsx, dashboard_data.json)")


if __name__ == "__main__":
    main()
