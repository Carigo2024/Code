"""
Leitor resiliente do relatório do MRP do Protheus (aba "Resultados").

Premissas validadas empiricamente (745 itens, arquivo de referência 2026-08):
  - O relatório é impresso em blocos repetidos, um por item.
  - Cada bloco começa numa linha cujo texto na coluna A é "Produto".
  - A linha logo abaixo é a "linha mestre" com os atributos estáticos.
  - As séries mensais (Estoque, Entradas, Saídas, Saldo Final, Necessidade...)
    ficam em linhas de detalhamento, cada uma rotulada na coluna A e alinhada
    às datas da linha "Período".

O leitor NÃO assume posições fixas de coluna nem quantidade fixa de meses ou
de itens: detecta o período pela linha "Período", as colunas de atributo pelos
rótulos do cabeçalho e cada série pelo texto do rótulo. Se a estrutura mínima
não for encontrada, levanta ErroEstruturaMRP em vez de devolver dados errados.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime

import openpyxl

from . import config


class ErroEstruturaMRP(Exception):
    """Estrutura do relatório do MRP não bate com o esperado."""


@dataclass
class ItemMRP:
    codigo: str
    descricao: str
    tipo: str
    lead_time: float
    armazem: str
    lote_minimo: float
    lote_economico: float
    estoque_seguranca: float
    ponto_pedido: float
    unidade: str
    valor_unitario: float
    # Séries mensais (uma lista por rótulo de detalhamento), alinhadas a `periodos`
    series: dict[str, list[float]] = field(default_factory=dict)

    def serie(self, rotulo: str, n: int) -> list[float]:
        """Série mensal do rótulo, ou zeros se ausente (com tamanho n)."""
        return self.series.get(rotulo, [0.0] * n)


def _num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(".", "").replace(",", ".")) if "," in str(v) else float(v)
    except (ValueError, TypeError):
        return 0.0


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _achar_linha_periodo(rows: list[tuple]) -> int:
    for i, r in enumerate(rows):
        if r and _txt(r[0]).lower().startswith(config.ROTULO_PERIODO.lower()):
            return i
    raise ErroEstruturaMRP(
        f"Linha '{config.ROTULO_PERIODO}' não encontrada nas primeiras linhas do relatório."
    )


def _ler_periodos(linha: tuple) -> list[datetime]:
    """Lê as datas mensais contíguas à direita do rótulo 'Período'."""
    periodos: list[datetime] = []
    for v in linha[1:]:
        if isinstance(v, datetime):
            periodos.append(v)
        elif v is None or _txt(v) == "":
            if periodos:  # parou de encontrar datas -> fim do bloco de meses
                break
        else:
            # tenta interpretar dd/mm/aaaa
            try:
                periodos.append(datetime.strptime(_txt(v), "%d/%m/%Y"))
            except ValueError:
                if periodos:
                    break
    if not periodos:
        raise ErroEstruturaMRP("Nenhuma data mensal encontrada na linha 'Período'.")
    return periodos


def _mapear_cabecalho(linha_produto: tuple) -> dict[str, int]:
    """Mapeia rótulo do cabeçalho -> índice de coluna, tolerante a acentos/caixa."""
    mapa: dict[str, int] = {}
    for idx, v in enumerate(linha_produto):
        t = _txt(v).lower()
        if t:
            mapa[t] = idx
    return mapa


def _col(mapa: dict[str, int], *nomes: str, obrigatorio: bool = True) -> int | None:
    for nome in nomes:
        if nome.lower() in mapa:
            return mapa[nome.lower()]
    if obrigatorio:
        raise ErroEstruturaMRP(
            f"Coluna de atributo não encontrada no cabeçalho (procurado: {nomes})."
        )
    return None


def ler_relatorio_mrp(caminho: str, aba: str = "Resultados") -> tuple[list[ItemMRP], list[datetime], list[str]]:
    """
    Lê o relatório do MRP e devolve (itens, periodos, avisos).

    `avisos` traz problemas não-fatais detectados (ex.: bloco sem linha mestre),
    para o usuário decidir se confia no resultado.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        raise ErroEstruturaMRP(
            f"Aba '{aba}' não existe no arquivo. Abas disponíveis: {wb.sheetnames}"
        )
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ErroEstruturaMRP("Aba de resultados vazia.")

    avisos: list[str] = []

    # 1) Período (define quantos meses e as datas)
    lin_per = _achar_linha_periodo(rows)
    periodos = _ler_periodos(rows[lin_per])
    n = len(periodos)

    # 2) Localiza blocos pela ocorrência de "Produto" na coluna A
    starts = [i for i, r in enumerate(rows) if r and _txt(r[0]) == config.ROTULO_PRODUTO]
    if not starts:
        raise ErroEstruturaMRP(
            f"Nenhum bloco de item encontrado (nenhuma linha '{config.ROTULO_PRODUTO}' na coluna A)."
        )

    # 3) Cabeçalho de atributos: primeira linha "Produto"
    mapa = _mapear_cabecalho(rows[starts[0]])
    c_desc = _col(mapa, "descrição", "descricao")
    c_tipo = _col(mapa, "tipo")
    c_lead = _col(mapa, "lead time")
    c_arm = _col(mapa, "armazém", "armazem")
    c_lmin = _col(mapa, "lote mínimo", "lote minimo")
    c_lecon = _col(mapa, "lote econômico", "lote economico")
    c_eseg = _col(mapa, "estoque segurança", "estoque seguranca")
    c_pp = _col(mapa, "ponto pedido")
    c_um = _col(mapa, "unidade de medida")
    c_val = _col(mapa, "valor")

    rotulos_detalhe = {
        config.ROTULO_ESTOQUE, config.ROTULO_ENTRADAS, config.ROTULO_SAIDAS,
        config.ROTULO_SAIDA_ESTRUTURA, config.ROTULO_TRANSF_SAIDA,
        config.ROTULO_TRANSF_ENTRADA, config.ROTULO_SALDO_FINAL,
        config.ROTULO_NECESSIDADE,
    }

    itens: list[ItemMRP] = []
    for bi, s in enumerate(starts):
        fim = starts[bi + 1] if bi + 1 < len(starts) else len(rows)
        mestre = rows[s + 1] if s + 1 < fim else None
        if mestre is None or _txt(mestre[0]) == "":
            avisos.append(f"Bloco na linha {s + 1}: sem linha mestre; item ignorado.")
            continue

        item = ItemMRP(
            codigo=_txt(mestre[0]),
            descricao=_txt(mestre[c_desc]),
            tipo=_txt(mestre[c_tipo]),
            lead_time=_num(mestre[c_lead]),
            armazem=_txt(mestre[c_arm]).zfill(2),
            lote_minimo=_num(mestre[c_lmin]),
            lote_economico=_num(mestre[c_lecon]),
            estoque_seguranca=_num(mestre[c_eseg]),
            ponto_pedido=_num(mestre[c_pp]),
            unidade=_txt(mestre[c_um]),
            valor_unitario=_num(mestre[c_val]),
        )

        # séries mensais: linhas de detalhe até o próximo bloco
        for r in range(s + 2, fim):
            rot = _txt(rows[r][0])
            if rot in rotulos_detalhe and rot not in item.series:
                item.series[rot] = [_num(rows[r][1 + k]) for k in range(n)]
        itens.append(item)

    # 4) Validação de sanidade: a identidade do Saldo Final deve fechar
    _validar_identidade_saldo(itens, n, avisos)
    wb.close()
    return itens, periodos, avisos


def _validar_identidade_saldo(itens: list[ItemMRP], n: int, avisos: list[str]) -> None:
    """
    Confirma Saldo Final = +Estoque + Entradas - (componentes de saída).
    Se muitos itens não fecharem, a estrutura provavelmente mudou -> avisa.
    """
    falhas = 0
    checados = 0
    for it in itens:
        if config.ROTULO_SALDO_FINAL not in it.series:
            continue
        est = it.serie(config.ROTULO_ESTOQUE, n)
        ent = it.serie(config.ROTULO_ENTRADAS, n)
        saldo = it.serie(config.ROTULO_SALDO_FINAL, n)
        checados += 1
        for m in range(n):
            saida = sum(sinal * it.serie(rot, n)[m]
                        for rot, sinal in config.COMPONENTES_SAIDA.items())
            calc = est[m] + ent[m] - saida
            if abs(calc - saldo[m]) > 0.5:
                falhas += 1
                break
    if checados and falhas / checados > 0.05:
        avisos.append(
            f"ATENÇÃO: identidade do Saldo Final falhou em {falhas}/{checados} itens "
            "(>5%). A estrutura do relatório pode ter mudado — revise antes de confiar."
        )
