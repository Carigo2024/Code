"""
Exportação do resultado para Excel.

ATENÇÃO: o layout aqui é PRELIMINAR. O formato definitivo (abas, colunas,
formatação) deve seguir o arquivo-modelo que o usuário ainda vai fornecer.
Este exportador expõe TODOS os valores intermediários (estoque-base, mês de
ruptura, saldo mínimo, cobertura, Necessidade Protheus, alertas) para que
qualquer decisão seja auditável. Ao receber o modelo, só a montagem das
colunas/abas muda — o motor não.
"""
from __future__ import annotations

from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .motor import Decisao

_VERMELHO = PatternFill("solid", fgColor="F4CCCC")   # urgente / revisão manual
_AMARELO = PatternFill("solid", fgColor="FFF2CC")    # divergente do Protheus
_CINZA = PatternFill("solid", fgColor="D9D9D9")      # cabeçalho
_HDR_FONT = Font(bold=True)

COLUNAS = [
    ("Código", "codigo"),
    ("Descrição", "descricao"),
    ("Tipo", "tipo"),
    ("Natureza", "natureza"),
    ("Ação", "acao"),
    ("Urgente", lambda d: "SIM" if d.urgente else ""),
    ("Volume", "volume"),
    ("Estoque disp. (01/05/22)", "estoque_base"),
    ("Estoque MRP m1", "estoque_mrp_m1"),
    ("Diverg. estoque", "divergencia_estoque"),
    ("Estoque Segurança", "estoque_seguranca"),
    ("Lead time (d)", "lead_time"),
    ("Lote Mín", "lote_minimo"),
    ("Lote Econ", "lote_economico"),
    ("Mês ruptura", lambda d: d.data_ruptura.strftime("%m/%Y") if d.data_ruptura else ""),
    ("Saldo mínimo", "saldo_min"),
    ("Cobertura (d)", lambda d: "" if d.cobertura_dias == float("inf") else round(d.cobertura_dias)),
    ("Necessidade Protheus", "necessidade_protheus_total"),
    ("Diverge Protheus", lambda d: "SIM" if d.divergente_do_protheus else ""),
    ("Revisão manual", lambda d: "SIM" if d.revisao_manual else ""),
    ("Alertas", lambda d: " | ".join(d.alertas)),
    ("Justificativa", "justificativa"),
]


def _valor(d: Decisao, chave):
    return chave(d) if callable(chave) else getattr(d, chave)


def exportar(decisoes: list[Decisao], caminho: str, periodos: list[datetime]) -> None:
    wb = openpyxl.Workbook()

    # --- aba principal ---
    ws = wb.active
    ws.title = "Analise"
    ws.append([c[0] for c in COLUNAS])
    for cell in ws[1]:
        cell.font = _HDR_FONT
        cell.fill = _CINZA
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    # ordena por urgência e depois por volume desc, revisão manual no topo
    ordenadas = sorted(
        decisoes,
        key=lambda d: (not d.revisao_manual, not d.urgente, not d.divergente_do_protheus, -d.volume),
    )
    for d in ordenadas:
        ws.append([_valor(d, c[1]) for c in COLUNAS])
        linha = ws[ws.max_row]
        if d.revisao_manual or d.urgente:
            for cell in linha:
                cell.fill = _VERMELHO
        elif d.divergente_do_protheus:
            for cell in linha:
                cell.fill = _AMARELO

    larguras = [12, 34, 6, 10, 18, 8, 12, 16, 14, 13, 14, 11, 10, 10, 11, 12, 11, 16, 12, 12, 40, 80]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- aba resumo ---
    rs = wb.create_sheet("Resumo")
    total = len(decisoes)
    comprar = sum(1 for d in decisoes if d.natureza == "Comprar" and d.volume > 0)
    produzir = sum(1 for d in decisoes if d.natureza == "Produzir" and d.volume > 0)
    ok = sum(1 for d in decisoes if d.mes_ruptura_idx is None and not d.revisao_manual)
    urg = sum(1 for d in decisoes if d.urgente)
    div = sum(1 for d in decisoes if d.divergente_do_protheus)
    rev = sum(1 for d in decisoes if d.revisao_manual)
    recon = sum(1 for d in decisoes if any("diverge do +Estoque" in a for a in d.alertas))
    resumo = [
        ("Relatório gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Horizonte", f"{periodos[0]:%m/%Y} a {periodos[-1]:%m/%Y} ({len(periodos)} meses)"),
        ("Itens analisados", total),
        ("Comprar (com volume)", comprar),
        ("Produzir (com volume)", produzir),
        ("Coberto (sem pedido)", ok),
        ("Urgentes", urg),
        ("Divergentes do Protheus", div),
        ("Alerta reconciliação estoque", recon),
        ("Revisão manual (importados)", rev),
    ]
    for k, v in resumo:
        rs.append([k, v])
    rs.column_dimensions["A"].width = 34
    rs.column_dimensions["B"].width = 40
    for cell in rs["A"]:
        cell.font = _HDR_FONT

    wb.save(caminho)
