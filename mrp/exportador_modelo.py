"""
Exportação no LAYOUT do arquivo-modelo `Order_MM_AAAA` (aba "FORECAST FINAL").

O modelo é uma matriz por componente importado:
  A = COMPONENT UDB | B = COMPONENT UPI | C = DESCRIPTION | D = UM
  E..P = 12 meses sob "MONTH OF AVAILABILITY" | Q = marcador (Tipo)

Os valores mensais aqui são a NECESSIDADE MENSAL time-phased calculada pelo
motor (quanto precisa ficar disponível em cada mês para manter o saldo acima do
estoque de segurança) — derivada do MRP, NÃO do forecast externo que preencheu
o modelo original. A semântica exata (escopo de itens, fonte do COMPONENT UPI)
ainda está em confirmação com o usuário; este exportador é um rascunho fiel ao
formato, fácil de ajustar quando as respostas chegarem.
"""
from __future__ import annotations

from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import config
from .motor import Decisao

_HDR_FONT = Font(bold=True)
ABA = "FORECAST FINAL (ORIG)"


def _eh_importado(d: Decisao) -> bool:
    """Escopo do modelo: importados. IM + itens de contrato (-BR) e revisão manual."""
    return (
        d.tipo == "IM"
        or d.codigo in config.SKUS_REVISAO_MANUAL
        or d.codigo.upper().endswith("-BR")
    )


def exportar_modelo(
    decisoes: list[Decisao],
    caminho: str,
    periodos: list[datetime],
    apenas_importados: bool = True,
) -> int:
    n = len(periodos)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ABA

    # --- linha 1: "MONTH OF AVAILABILITY" mesclado sobre os meses ---
    col_ini_mes = 5  # coluna E
    ws.cell(row=1, column=col_ini_mes, value="MONTH OF AVAILABILITY").font = _HDR_FONT
    ws.merge_cells(
        start_row=1, start_column=col_ini_mes,
        end_row=1, end_column=col_ini_mes + n - 1,
    )

    # --- linha 2: cabeçalho ---
    ws.cell(row=2, column=1, value="COMPONENT UDB")
    ws.cell(row=2, column=2, value="COMPONENT UPI")
    ws.cell(row=2, column=3, value="DESCRIPTION")
    ws.cell(row=2, column=4, value="UM")
    for k, per in enumerate(periodos):
        c = ws.cell(row=2, column=col_ini_mes + k, value=per)
        c.number_format = "mm/yyyy"
    col_tipo = col_ini_mes + n
    ws.cell(row=2, column=col_tipo, value="TIPO")
    for cell in ws[2]:
        cell.font = _HDR_FONT
        cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = ws.cell(row=3, column=4).coordinate  # congela A-C e linhas 1-2

    # --- dados ---
    alvo = [d for d in decisoes if (not apenas_importados or _eh_importado(d))]
    # ordena: revisão manual no topo, depois maior necessidade total
    alvo.sort(key=lambda d: (not d.revisao_manual, -sum(d.necessidade_mensal)))

    linha = 3
    for d in alvo:
        ws.cell(row=linha, column=1, value=d.codigo)
        ws.cell(row=linha, column=2, value=f"{d.codigo}-")  # COMPONENT UPI: PENDENTE fonte real
        ws.cell(row=linha, column=3, value=d.descricao)
        ws.cell(row=linha, column=4, value="UN")
        for k in range(n):
            v = d.necessidade_mensal[k] if k < len(d.necessidade_mensal) else 0.0
            cc = ws.cell(row=linha, column=col_ini_mes + k, value=round(v))
            cc.number_format = "#,##0"
        ws.cell(row=linha, column=col_tipo, value=d.tipo)
        linha += 1

    larguras = [27.9, 27.0, 58.5, 11.7] + [18.0] * n + [10.6]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(caminho)
    return len(alvo)
