"""
Leitor do arquivo de estoque (aba "Listagem do Browse"), nível de lote.

Regra de disponibilidade confirmada com o usuário:
  - disponível por SKU = Σ (Quantidade - Empenho), pois o Empenho (reservado)
    é considerado JÁ CONSUMIDO;
  - somente os armazéns de config.ARMAZENS_ALVO (01, 05, 22) entram na conta.

O arquivo é uma tabela plana com o cabeçalho na 2ª linha (a 1ª é um título).
Também sinaliza quantidade que já passou do Pull Date (não deve ser consumida),
sem removê-la do total — apenas registra para alerta.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime

import openpyxl

from . import config


class ErroEstruturaEstoque(Exception):
    """Estrutura do arquivo de estoque não bate com o esperado."""


@dataclass
class EstoqueSKU:
    codigo: str
    disponivel: float = 0.0          # Σ(Qtde-Empenho) nos armazéns-alvo
    quantidade_bruta: float = 0.0    # Σ Qtde nos armazéns-alvo (sem tirar empenho)
    empenho: float = 0.0             # Σ Empenho nos armazéns-alvo
    qtd_apos_pull_date: float = 0.0  # quantidade em lotes já vencidos p/ consumo
    n_lotes: int = 0
    armazens: set[str] = field(default_factory=set)


def _num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _achar_cabecalho(rows: list[tuple]) -> int:
    """Acha a linha de cabeçalho (a que contém 'Produto' e 'Quantidade')."""
    for i, r in enumerate(rows[:10]):
        vals = {_txt(x).lower() for x in r}
        if "produto" in vals and "quantidade" in vals:
            return i
    raise ErroEstruturaEstoque(
        "Cabeçalho não encontrado (nenhuma linha com 'Produto' e 'Quantidade')."
    )


def ler_estoque(
    caminho: str,
    aba: str = "Listagem do Browse",
    armazens_alvo: set[str] | None = None,
    data_referencia: datetime | None = None,
) -> tuple[dict[str, EstoqueSKU], list[str]]:
    """
    Lê o arquivo de estoque e devolve (estoque_por_sku, avisos).
    `armazens_alvo` default = config.ARMAZENS_ALVO.
    """
    armazens_alvo = armazens_alvo or config.ARMAZENS_ALVO
    data_referencia = data_referencia or datetime.now()

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    if aba not in wb.sheetnames:
        raise ErroEstruturaEstoque(
            f"Aba '{aba}' não existe. Abas disponíveis: {wb.sheetnames}"
        )
    ws = wb[aba]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ErroEstruturaEstoque("Aba de estoque vazia.")

    h = _achar_cabecalho(rows)
    hdr = {_txt(v).lower(): i for i, v in enumerate(rows[h]) if _txt(v)}

    def col(*nomes: str) -> int:
        for nome in nomes:
            if nome.lower() in hdr:
                return hdr[nome.lower()]
        raise ErroEstruturaEstoque(f"Coluna não encontrada no estoque: {nomes}")

    c_prod = col("produto")
    c_arm = col("armazem", "armazém")
    c_qtd = col("quantidade")
    c_emp = col("empenho")
    c_pull = hdr.get("pull date")  # opcional

    avisos: list[str] = []
    estoque: dict[str, EstoqueSKU] = {}
    armazens_vistos: set[str] = set()

    for r in rows[h + 1:]:
        if not r or r[c_prod] is None or _txt(r[c_prod]) == "":
            continue
        arm = _txt(r[c_arm]).zfill(2)
        armazens_vistos.add(arm)
        if arm not in armazens_alvo:
            continue
        cod = _txt(r[c_prod])
        e = estoque.setdefault(cod, EstoqueSKU(codigo=cod))
        q = _num(r[c_qtd])
        emp = _num(r[c_emp])
        e.quantidade_bruta += q
        e.empenho += emp
        e.disponivel += q - emp
        e.n_lotes += 1
        e.armazens.add(arm)
        if c_pull is not None and isinstance(r[c_pull], datetime) and r[c_pull] < data_referencia:
            e.qtd_apos_pull_date += max(q - emp, 0.0)

    if not estoque:
        avisos.append(
            f"Nenhum lote encontrado nos armazéns-alvo {sorted(armazens_alvo)}. "
            f"Armazéns presentes no arquivo: {sorted(armazens_vistos)}."
        )
    com_pull = sum(1 for e in estoque.values() if e.qtd_apos_pull_date > 0)
    if com_pull:
        avisos.append(
            f"{com_pull} SKU(s) têm quantidade além do Pull Date nos armazéns-alvo — "
            "não devem ser consumidos sem revisão."
        )
    wb.close()
    return estoque, avisos
