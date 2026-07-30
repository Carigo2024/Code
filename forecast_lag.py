"""
Acuracidade de forecast de importacao por LAG (defasagem da safra).

Grao do modelo
--------------
    Item  x  Mes-alvo  x  Safra do forecast (mes em que a previsao foi gerada)

Desse grao deriva o LAG = mes-alvo - safra (em meses):
    Lag 0  = previsao feita no proprio mes-alvo
    Lag 1  = feita 1 mes antes
    ...    ate Lag 11.

Dois modos de leitura
---------------------
    * LAG FIXO (honesto): compara sempre a previsao feita com N meses de
      antecedencia. N e proposto a partir do lead time de importacao real
      (mediana de Invoice Date - Purchase Order Date no arquivo de recebidos).
    * CONSOLIDADO (gerencial): compara com a previsao mais recente que cobre
      cada mes-alvo. Leitura otimista - usa informacao que so existia depois
      de a compra ja ter sido travada.

Tratamento da defasagem fisica do "recebido"
--------------------------------------------
O recebido e DATA DE ENTRADA FISICA, nao demanda. O mesmo erro de volume
aparece inflado ou deflacionado conforme o timing de chegada. Por isso as
metricas sao calculadas em tres leituras:
    (a) mes a mes      - sensivel a timing (mais dura, injusta com o forecast)
    (b) buckets trim.  - absorve deslizes de +-1 mes dentro do trimestre
    (c) acumulada      - neutraliza timing e isola o erro de VOLUME

Uso
---
    python forecast_lag.py
    python forecast_lag.py --data-dir data --lag 2 --output forecast_lag.png
    python forecast_lag.py --csv grao_forecast.csv        # exporta o grao tidy

Premissa de datacao da safra (documentada, ajustavel via --safras)
------------------------------------------------------------------
Os metadados dos arquivos foram sobrescritos por re-gravacoes, entao a safra
e datada pelo numero do arquivo: Order_0N -> mes N de 2026 (Jan, Fev, Mar,
Abr). Isso preserva o grao "uma safra por mes". O Order_03 comeca em Abr/26
(nao cobre Mar) e traz uma coluna Fev/27 duplicada - ambos tratados no parser.
Datar a safra pela primeira coluna colapsaria Order_03 e Order_04 na mesma
safra (Abr), o que o grao proibe; ver README_forecast_lag.md.
"""

import argparse
import glob
import os
import re
import sys

import numpy as np
import pandas as pd

# --- Paleta validada (CVD-safe, ver skill dataviz) ---
COR_FCST = "#2a78d6"   # azul  - previsao
COR_RECV = "#eb6834"   # laranja - recebido
COR_POS = "#e34948"    # vermelho - recebeu MAIS que o previsto
COR_NEG = "#2a78d6"    # azul    - recebeu MENOS que o previsto
COR_TINTA = "#52514e"  # texto secundario
COR_GRID = "#e7e6e2"

# Mapeamento padrao arquivo -> safra (ano, mes). Order_0N = mes N de 2026.
SAFRAS_PADRAO = {1: (2026, 1), 2: (2026, 2), 3: (2026, 3), 4: (2026, 4)}


# --------------------------------------------------------------------------- #
# Utilidades de tempo (mes como inteiro continuo para aritmetica de lag)
# --------------------------------------------------------------------------- #
def chave_mes(ano, mes):
    """Converte (ano, mes) num inteiro continuo: ano*12 + (mes-1)."""
    return ano * 12 + (mes - 1)


def rotulo_mes(k):
    return f"{k // 12:04d}-{k % 12 + 1:02d}"


def rotulo_trim(k):
    ano, mes = k // 12, k % 12 + 1
    return f"{ano}Q{(mes - 1) // 3 + 1}"


def normaliza_item(x):
    """Normaliza o codigo do item para casar forecast x recebido.

    Remove sufixo '-BR' e traco final (COMPONENT UPI vem como '12976-').
    """
    s = str(x).strip().upper()
    s = re.sub(r"-BR$", "", s)
    s = re.sub(r"-$", "", s)
    return s


# --------------------------------------------------------------------------- #
# Leitura dos arquivos
# --------------------------------------------------------------------------- #
def descobre_arquivo_order(data_dir, n):
    """Acha o Order_0N no diretorio (aceita qualquer sufixo de nome)."""
    padroes = [f"*Order_{n:02d}_*.xlsx", f"*Order_{n}_*.xlsx", f"*Order_{n:02d}*.xlsx"]
    for p in padroes:
        achados = sorted(glob.glob(os.path.join(data_dir, p)))
        if achados:
            return achados[0]
    return None


def le_forecast(caminho, safra_ano, safra_mes):
    """Le uma safra de forecast e devolve linhas (item, alvo, safra, lag, fcst).

    Layout: linha 1 titulo, linha 2 cabecalho (col A-D metadados, col E+ meses),
    dados a partir da linha 3. Colunas de mes repetidas sao deduplicadas
    mantendo a primeira ocorrencia.
    """
    import openpyxl
    import datetime as dt

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["FORECAST"]
    linhas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    cabecalho = linhas[0]
    safra = chave_mes(safra_ano, safra_mes)
    vistos = set()
    colunas = []  # (indice_coluna, chave_mes_alvo)
    for j in range(4, len(cabecalho)):
        h = cabecalho[j]
        if isinstance(h, dt.datetime):
            tk = chave_mes(h.year, h.month)
            if tk in vistos:
                continue
            vistos.add(tk)
            colunas.append((j, tk))

    recs = []
    for r in linhas[1:]:
        item = r[0]
        if item in (None, "") or not str(item).strip():
            continue
        item = normaliza_item(item)
        for j, tk in colunas:
            q = r[j]
            if q is None:
                continue
            recs.append((item, tk, safra, tk - safra, float(q)))
    return recs


def carrega_grao(data_dir, safras):
    """Monta o grao tidy do forecast: Item x Mes-alvo x Safra -> lag, fcst."""
    recs = []
    for n, (ano, mes) in sorted(safras.items()):
        caminho = descobre_arquivo_order(data_dir, n)
        if not caminho:
            print(f"  aviso: Order_{n:02d} nao encontrado em {data_dir}", file=sys.stderr)
            continue
        novos = le_forecast(caminho, ano, mes)
        recs.extend(novos)
        print(f"  safra {ano}-{mes:02d}  <- {os.path.basename(caminho)}  ({len(novos)} celulas)")
    return pd.DataFrame(recs, columns=["item", "alvo", "safra", "lag", "fcst"])


def carrega_recebido(data_dir):
    """Le o recebido e agrega por (item, mes de entrada fisica).

    Usa 'Invoice Date' como mes-alvo e 'Qtd Delivered' como quantidade real.
    Tambem devolve a serie de lead time (Invoice - PO) em dias.
    """
    import openpyxl
    import datetime as dt

    achados = sorted(glob.glob(os.path.join(data_dir, "*Received*.xlsx")))
    if not achados:
        raise FileNotFoundError(f"arquivo Received*.xlsx nao encontrado em {data_dir}")
    caminho = achados[0]
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["Export"]
    linhas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    print(f"  recebido    <- {os.path.basename(caminho)}  ({len(linhas)} linhas)")

    recs, leads = [], []
    for r in linhas:
        # 0 Invoice 1 InvoiceDate 2 PO 3 PODate 4 Supplier 5 Item ... 12 QtdDelivered
        if r[5] in (None, ""):
            continue
        d, q = r[1], r[12]
        if not isinstance(d, dt.datetime) or q is None:
            continue
        recs.append((normaliza_item(r[5]), chave_mes(d.year, d.month), float(q)))
        if isinstance(r[3], dt.datetime):
            leads.append((d - r[3]).days)

    act = (
        pd.DataFrame(recs, columns=["item", "alvo", "recv"])
        .groupby(["item", "alvo"], as_index=False)
        .recv.sum()
    )
    return act, np.array(leads)


# --------------------------------------------------------------------------- #
# Selecao das previsoes (os dois modos de leitura)
# --------------------------------------------------------------------------- #
def modo_lag_fixo(fc, n):
    """Previsao feita exatamente com N meses de antecedencia."""
    return fc[fc.lag == n][["item", "alvo", "fcst"]]


def modo_consolidado(fc):
    """Previsao mais recente (maior safra) que cobre cada item x mes-alvo."""
    return (
        fc.sort_values("safra")
        .groupby(["item", "alvo"])
        .tail(1)[["item", "alvo", "fcst"]]
    )


# --------------------------------------------------------------------------- #
# Metricas nas tres leituras de timing
# --------------------------------------------------------------------------- #
def _wape(dif_abs, denom):
    return float(dif_abs.sum() / denom) if denom else float("nan")


def metricas(sel, act, janela):
    """Calcula bias e WAPE nas tres leituras para uma janela de meses-alvo.

    Retorna dict com totais e WAPE mensal / trimestral / acumulada. Itens sem
    par (previu e nao recebeu, ou vice-versa) entram como zero do outro lado.
    """
    m = sel.merge(act, on=["item", "alvo"], how="outer")
    m = m[m.alvo.isin(janela)].copy()
    m[["fcst", "recv"]] = m[["fcst", "recv"]].fillna(0.0)

    F, A = m.fcst.sum(), m.recv.sum()
    bias = (A - F) / F if F else float("nan")

    # (a) mes a mes: erro no grao item x mes
    wape_mes = _wape((m.fcst - m.recv).abs(), A)

    # (b) buckets trimestrais: agrega item x trimestre antes do erro
    mq = m.assign(trim=m.alvo.map(rotulo_trim))
    g = mq.groupby(["item", "trim"]).agg(f=("fcst", "sum"), a=("recv", "sum"))
    wape_trim = _wape((g.f - g.a).abs(), A)

    # (c) acumulada: total item na janela inteira (timing some totalmente)
    gc = m.groupby("item").agg(f=("fcst", "sum"), a=("recv", "sum"))
    wape_acum = _wape((gc.f - gc.a).abs(), A)

    return dict(F=F, A=A, bias=bias, wape_mes=wape_mes,
                wape_trim=wape_trim, wape_acum=wape_acum, n_itens=m.item.nunique())


def serie_mensal(sel, act, janela):
    """Totais por mes-alvo + acumulados, para tabela e grafico."""
    m = sel.merge(act, on=["item", "alvo"], how="outer")
    m = m[m.alvo.isin(janela)].copy()
    m[["fcst", "recv"]] = m[["fcst", "recv"]].fillna(0.0)
    s = (
        m.groupby("alvo").agg(F=("fcst", "sum"), A=("recv", "sum"))
        .reindex(sorted(janela)).reset_index()
    )
    s["cumF"] = s.F.cumsum()
    s["cumA"] = s.A.cumsum()
    return s


# --------------------------------------------------------------------------- #
# Escolha do N (lead time de importacao)
# --------------------------------------------------------------------------- #
def sugere_n(leads):
    """Propoe o lag fixo N = round(mediana do lead time em meses)."""
    if len(leads) == 0:
        return 2, float("nan")
    meses = np.median(leads) / 30.44
    return int(max(1, round(meses))), float(np.median(leads))


# --------------------------------------------------------------------------- #
# Grafico
# --------------------------------------------------------------------------- #
def plota(serie_lag, met_lag, met_cons_win, n, output):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    milhar = FuncFormatter(lambda v, _: f"{v/1e6:.1f}M" if abs(v) >= 1e6 else f"{v/1e3:.0f}k")
    plt.rcParams.update({"axes.edgecolor": COR_TINTA, "axes.labelcolor": COR_TINTA,
                         "xtick.color": COR_TINTA, "ytick.color": COR_TINTA,
                         "font.size": 10, "axes.spines.top": False, "axes.spines.right": False})

    fig, axs = plt.subplots(2, 2, figsize=(14, 9.5))
    fig.suptitle(f"Acuracidade do forecast de importacao  -  leitura honesta = LAG {n}",
                 fontsize=16, fontweight="bold")

    meses = [rotulo_mes(k) for k in serie_lag.alvo]
    x = np.arange(len(meses))

    # Painel 1: mes a mes, Previsto vs Recebido (barras agrupadas)
    ax = axs[0, 0]
    w = 0.4
    b1 = ax.bar(x - w/2, serie_lag.F, w, color=COR_FCST, label=f"Previsto (lag {n})")
    b2 = ax.bar(x + w/2, serie_lag.A, w, color=COR_RECV, label="Recebido (entrada fisica)")
    ax.set_title("(a) Mes a mes  -  inflada por timing de chegada", fontweight="bold")
    ax.set_xticks(x); ax.set_xticklabels(meses)
    ax.yaxis.set_major_formatter(milhar); ax.grid(axis="y", color=COR_GRID)
    ax.legend(frameon=False, fontsize=9)
    for _, r in serie_lag.iterrows():
        i = list(serie_lag.alvo).index(r.alvo)
        err = (r.A - r.F) / r.F * 100 if r.F else 0
        ax.annotate(f"{err:+.0f}%", (i, max(r.F, r.A)), ha="center", va="bottom",
                    fontsize=8, color=COR_POS if err >= 0 else COR_NEG, fontweight="bold")

    # Painel 2: acumulada (linhas convergindo)
    ax = axs[0, 1]
    ax.plot(x, serie_lag.cumF, "-o", color=COR_FCST, lw=2, ms=6, label="Previsto acumulado")
    ax.plot(x, serie_lag.cumA, "-o", color=COR_RECV, lw=2, ms=6, label="Recebido acumulado")
    ax.fill_between(x, serie_lag.cumF, serie_lag.cumA, color=COR_RECV, alpha=0.08)
    ax.set_title("(c) Acumulada  -  neutraliza timing, isola VOLUME", fontweight="bold")
    ax.set_xticks(x); ax.set_xticklabels(meses)
    ax.yaxis.set_major_formatter(milhar); ax.grid(axis="y", color=COR_GRID)
    ax.legend(frameon=False, fontsize=9)
    biasf = met_lag["bias"] * 100
    ax.annotate(f"gap de volume: {biasf:+.1f}%", (x[-1], serie_lag.cumA.iloc[-1]),
                ha="right", va="bottom", fontsize=10, fontweight="bold",
                color=COR_POS if biasf >= 0 else COR_NEG)

    # Painel 3: WAPE nas tres leituras (o colapso do timing)
    ax = axs[1, 0]
    vals = [met_lag["wape_mes"]*100, met_lag["wape_trim"]*100, met_lag["wape_acum"]*100]
    cols = ["#c9c8c3", "#8fb8e6", COR_FCST]
    labels3 = ["(a) mes a mes\ntiming (injusto)", "(b) trimestral\nmeio-termo",
               "(c) acumulada\nvolume (justo)"]
    bars = ax.bar(labels3, vals, color=cols, edgecolor="white")
    ax.bar_label(bars, fmt="%.0f%%", fontweight="bold", padding=2)
    ax.set_title("Erro (WAPE) por leitura de timing  -  lag fixo", fontweight="bold")
    ax.set_ylabel("WAPE"); ax.grid(axis="y", color=COR_GRID)
    ax.set_ylim(0, max(vals) * 1.25)

    # Painel 4: contraste dos dois modos na MESMA janela
    ax = axs[1, 1]
    modos = [f"Lag {n}\n(honesto)", "Consolidado\n(gerencial)"]
    bias_vals = [met_lag["bias"]*100, met_cons_win["bias"]*100]
    acum_vals = [met_lag["wape_acum"]*100, met_cons_win["wape_acum"]*100]
    xx = np.arange(2)
    bA = ax.bar(xx - 0.2, np.abs(bias_vals), 0.4, color=COR_RECV, label="|Bias| volume")
    bB = ax.bar(xx + 0.2, acum_vals, 0.4, color=COR_FCST, label="WAPE acumulada")
    ax.set_title("Dois modos na mesma janela-alvo", fontweight="bold")
    ax.set_xticks(xx); ax.set_xticklabels(modos)
    ax.set_ylabel("%"); ax.grid(axis="y", color=COR_GRID)
    ax.bar_label(bA, fmt="%.0f%%", fontsize=8); ax.bar_label(bB, fmt="%.0f%%", fontsize=8)
    ax.legend(frameon=False, fontsize=9)

    fig.tight_layout(rect=[0, 0, 1, 0.96])
    fig.savefig(output, dpi=130, bbox_inches="tight")
    print(f"\nGrafico salvo em: {output}")


# --------------------------------------------------------------------------- #
def imprime_bloco(titulo, met, janela):
    print(f"\n[{titulo}]")
    print(f"   janela-alvo: {', '.join(rotulo_mes(t) for t in sorted(janela))}  "
          f"({met['n_itens']} itens)")
    print(f"   Sigma Previsto = {met['F']:>13,.0f}   Sigma Recebido = {met['A']:>13,.0f}")
    print(f"   Bias de volume (Recebido/Previsto - 1) = {met['bias']*100:+.1f}%")
    print(f"   WAPE (a) mes a mes    = {met['wape_mes']*100:6.1f}%")
    print(f"   WAPE (b) trimestral   = {met['wape_trim']*100:6.1f}%")
    print(f"   WAPE (c) acumulada    = {met['wape_acum']*100:6.1f}%   <- isola volume")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--data-dir", default="data", help="diretorio com Order_0N e Received*.xlsx")
    ap.add_argument("--lag", type=int, default=None,
                    help="N do lag fixo (padrao: derivado do lead time de importacao)")
    ap.add_argument("--output", default="forecast_lag.png", help="PNG de saida")
    ap.add_argument("--csv", default=None, help="exporta o grao tidy para CSV")
    ap.add_argument("--no-plot", action="store_true", help="nao gerar grafico")
    args = ap.parse_args()

    print("Carregando forecast (safras):")
    fc = carrega_grao(args.data_dir, SAFRAS_PADRAO)
    print("Carregando recebido:")
    act, leads = carrega_recebido(args.data_dir)

    # escopo comparavel: itens presentes nos dois lados
    overlap = set(fc.item) & set(act.item)
    fc_o = fc[fc.item.isin(overlap)].copy()
    act_o = act[act.item.isin(overlap)].copy()
    print(f"\nItens no forecast: {fc.item.nunique()} | no recebido: {act.item.nunique()} "
          f"| comparaveis (interseccao): {len(overlap)}")

    if args.csv:
        fc.to_csv(args.csv, index=False)
        print(f"Grao tidy exportado: {args.csv}  ({len(fc)} linhas)")

    # N a partir do lead time
    n_auto, med_dias = sugere_n(leads)
    n = args.lag if args.lag is not None else n_auto
    print(f"\nLead time de importacao (Invoice - PO): mediana {med_dias:.0f} dias "
          f"(~{med_dias/30.44:.1f} meses)  ->  LAG FIXO proposto N = {n_auto}"
          + ("" if args.lag is None else f"  (forcado para {n})"))

    amax = act_o.alvo.max()  # ultimo mes com recebido

    # janelas
    jan_lag = sorted(t for t in set(modo_lag_fixo(fc_o, n).alvo) & set(act_o.alvo) if t <= amax)
    jan_full = sorted(t for t in set(act_o.alvo) if t <= amax)

    met_lag = metricas(modo_lag_fixo(fc_o, n), act_o, jan_lag)
    met_cons_win = metricas(modo_consolidado(fc_o), act_o, jan_lag)      # mesma janela
    met_cons_full = metricas(modo_consolidado(fc_o), act_o, jan_full)    # janela plena

    print("\n" + "=" * 74)
    print("MODO 1 - LAG FIXO (honesto)")
    print("=" * 74)
    imprime_bloco(f"Lag {n} - previsao travada {n} meses antes", met_lag, jan_lag)

    print("\n" + "=" * 74)
    print("MODO 2 - CONSOLIDADO (gerencial, otimista)")
    print("=" * 74)
    imprime_bloco("Consolidado na MESMA janela (comparavel ao lag fixo)", met_cons_win, jan_lag)
    imprime_bloco("Consolidado na janela plena (Jan->ultimo recebido)", met_cons_full, jan_full)

    # sensibilidade por lag
    print("\n" + "=" * 74)
    print("SENSIBILIDADE POR LAG (cada lag na sua janela disponivel)")
    print("=" * 74)
    print(f"{'lag':>3} {'meses-alvo':<40} {'bias':>8} {'WAPE_mes':>9} {'WAPE_acum':>10}")
    for L in range(0, 7):
        sel = modo_lag_fixo(fc_o, L)
        w = sorted(t for t in set(sel.alvo) & set(act_o.alvo) if t <= amax)
        if not w:
            continue
        mL = metricas(sel, act_o, w)
        janela_txt = ",".join(rotulo_mes(t) for t in w)
        print(f"{L:>3} {janela_txt:<40} {mL['bias']*100:>+7.1f}% "
              f"{mL['wape_mes']*100:>8.1f}% {mL['wape_acum']*100:>9.1f}%")

    # tabela mensal detalhada da leitura honesta
    serie = serie_mensal(modo_lag_fixo(fc_o, n), act_o, jan_lag)
    print("\n" + "=" * 74)
    print(f"DETALHE MENSAL x ACUMULADO - leitura honesta (lag {n})")
    print("=" * 74)
    print(f"{'mes':8} {'Previsto':>12} {'Recebido':>12} {'err_mes':>8} "
          f"{'cumPrev':>12} {'cumRecb':>12} {'err_acum':>9}")
    for _, r in serie.iterrows():
        em = (r.A - r.F) / r.F * 100 if r.F else float("nan")
        ea = (r.cumA - r.cumF) / r.cumF * 100 if r.cumF else float("nan")
        print(f"{rotulo_mes(int(r.alvo)):8} {r.F:12,.0f} {r.A:12,.0f} {em:>+7.1f}% "
              f"{r.cumF:12,.0f} {r.cumA:12,.0f} {ea:>+8.1f}%")

    print("\nRECOMENDACAO: para julgar VOLUME, use a leitura (c) acumulada - "
          "neutraliza o timing de entrada fisica. A (b) trimestral e o meio-termo\n"
          "operacional; a (a) mes a mes serve so como diagnostico (superestima o erro).")

    if not args.no_plot:
        plota(serie, met_lag, met_cons_win, n, args.output)


if __name__ == "__main__":
    main()
